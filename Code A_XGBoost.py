###############################################################################
#  MONTHLY  HYBRID  (SARIMA  +  XGBOOST-ON-RESIDUALS)  PIPELINE  – PYTHON EDITION
#
#  • 3 timing scenarios: Lag-0, Lag-12, ARIMAX
#  • Stage-1 rolling-origin grid-search for best SARIMA + exogenous subset
#  • Stage-2 XGBoost trained on one-step SARIMA residuals
#    → final hybrid forecast = SARIMA prediction + XGB residual correction
#  • Walk-forward back-test of the hybrid model
#  • Monte-Carlo GARCH fan of hybrid residuals
#  • PNG diagnostics embedded in a Times New Roman 12-pt Excel workbook
#
#  ── REQUIRED LIBS ──
#  pip install pandas numpy openpyxl xlsxwriter matplotlib pillow joblib tqdm
#              statsmodels arch xgboost
###############################################################################

from __future__ import annotations
import os, math, tempfile, warnings, itertools, cProfile, pstats, io, time
import numpy as np, pandas as pd
import matplotlib.pyplot as plt
from joblib import Parallel, delayed, cpu_count
from tqdm import tqdm
from arch import arch_model
import xgboost as xgb
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import random
from statsforecast.models import AutoARIMA

# Suppress all warnings and warning output from subprocesses
warnings.filterwarnings('ignore')
os.environ['PYTHONWARNINGS'] = 'ignore'

SEED = 123
random.seed(SEED)
np.random.seed(SEED)
os.environ["PYTHONHASHSEED"] = str(SEED)

# ─── PARAMS (EDIT HERE) ──────────────────────────────────────────────────────
IN_FILE      = "/Users/shanechiu/Downloads/Index_MT1.xlsx"
OUT_FILE     = "/Users/shanechiu/Desktop/2022_11M_HYBRID.xlsx"

WIN_SIZE     = 120       # rolling-origin grid-search window (shrink for fair grid search)
STEP_SIZE_RO = 5         # grid-search step
WIN_WF       = 241       # walk-forward training window
STEP_WF      = 1         # walk-forward step
LAG_12       = 12
MC_N         = 3_000
MC_SEED      = 123
N_CPU        = cpu_count()                  # full core count
N_JOBS       = N_CPU                       # use every core

# XGBoost deterministic base
XGB_BASE = dict(
    max_depth=3, eta=0.1, tree_method="hist", predictor="cpu_predictor",
    subsample=1, colsample_bytree=1, seed=SEED, deterministic_histogram=1,
    objective="reg:squarederror", verbosity=0
)

HIT_GRID = [
    dict(max_depth=d, eta=e)
    for d in [2,3,4] for e in [0.05,0.1,0.2]
]

def _best_xgb(X, y):
    if X.shape[1] == 0: return None  # Only return None if no features at all
    best, best_hit = None, -1
    from sklearn.model_selection import train_test_split
    Xtr, Xte, ytr, yte = train_test_split(X, y, test_size=0.3, random_state=SEED)
    for g in HIT_GRID:
        params = {**XGB_BASE, **g}
        mdl = xgb.train(params, xgb.DMatrix(Xtr, label=ytr))
        pred = np.sign(mdl.predict(xgb.DMatrix(Xte)))
        hit  = (pred == np.sign(yte)).mean()
        if hit > best_hit: best, best_hit = mdl, hit
    best.save_model("best_xgb.json")   # cached for future runs
    print("Chosen XGB hit-rate:", round(best_hit,3))
    return best

# ─── HELPERS ─────────────────────────────────────────────────────────────────
def _parse_cols(s: str) -> list[int]:
    return [] if pd.isna(s) or s in ("", "None") else list(map(int, s.split(",")))

def _to_num(col: pd.Series) -> pd.Series:
    # strips %, commas, and anything that is not 0-9 . + -
    return pd.to_numeric(
        col.astype(str).str.replace(r"[^\d\.\-+]", "", regex=True),
        errors="coerce"
    )

def _hit_rate(wf: pd.DataFrame) -> float:
    return wf["Test_Hit"].mean()

def _auto_sarima(y: np.ndarray, x: np.ndarray | None):
    mdl = AutoARIMA(season_length=12)
    mdl = mdl.fit(y, X=x if x is not None and x.size else None)
    return mdl

def _forecast_exo_future(X: np.ndarray, idx_t: int) -> np.ndarray | None:
    if X.size == 0: return None
    fut = []
    for j in range(X.shape[1]):
        try:
            mdl = AutoARIMA(season_length=12)
            fit = mdl.fit(X[:idx_t, j])
            fut.append(float(fit.predict(1)[0]))
        except Exception:
            fut.append(np.nan)
    return np.array(fut).reshape(1, -1)

def _roll_origin_metrics(y, X, scenario) -> tuple[float, float, float]:
    n = len(y)
    if n <= WIN_SIZE: return (np.nan, np.nan, np.nan)
    se = ae = naive_ae = hits = tot = 0
    for i in range(WIN_SIZE, n - 1, STEP_SIZE_RO):
        y_tr = y[i - WIN_SIZE:i]
        X_tr = X[i - WIN_SIZE:i]
        if   scenario == "lag0":   X_te = None if X_tr.size == 0 else X[i:i+1]
        elif scenario == "lag12":  X_te = None if X_tr.size == 0 or i + 1 - LAG_12 <= 0 else X_tr[-LAG_12:-LAG_12+1]
        else:                      X_te = None if X_tr.size == 0 else _forecast_exo_future(X_tr, X_tr.shape[0])
        use_x = X_te is not None and X_tr.size
        try:
            mdl = _auto_sarima(y_tr, X_tr if use_x else None)
            fc  = mdl.predict(1, X=X_te)[0] if use_x else mdl.predict(1)[0]
        except Exception:
            continue
        act = y[i]
        if not math.isnan(fc) and not math.isnan(act):
            se       += (fc - act) ** 2
            ae       += abs(fc - act)
            naive_ae += abs(y_tr[-1] - act)
            hits     += int(np.sign(fc) == np.sign(act))
            tot      += 1
    if tot == 0: return (np.nan, np.nan, np.nan)
    return round(math.sqrt(se / tot), 4), round(ae / naive_ae, 4), round(hits / tot, 4)

def _grid_search(y, X, scenario) -> pd.DataFrame:
    n_exo = X.shape[1]
    combos = range(2 ** n_exo)
    def _eval(mask):
        cols = [i for i in range(n_exo) if mask & (1 << i)]
        X_mat = X[:, cols] if cols else np.empty((len(y), 0))
        rmse, mase, hit = _roll_origin_metrics(y, X_mat, scenario)
        subset = ",".join(map(str, cols)) if cols else "None"
        return subset, rmse, mase, hit
    res = Parallel(n_jobs=N_JOBS)(
        delayed(_eval)(m) for m in tqdm(combos, desc=f"Grid-{scenario} (2^{n_exo})")
    )
    df = pd.DataFrame(res, columns=["ExoSubset", "RMSE", "MASE", "HitRate"])
    df = df.sort_values(["HitRate", "RMSE", "MASE"], ascending=[False, True, True])
    df.insert(0, "Rank", range(1, len(df) + 1))
    return df

def _predict_xgb(model, X_row):
    if model is None or X_row is None: return 0.0
    return float(model.predict(xgb.DMatrix(X_row))[0])

def _monte_carlo(model, exo_row):
    from numpy.random import default_rng
    rng = default_rng(MC_SEED)
    resids = model.residuals
    if len(resids) < 20:
        raise RuntimeError("Not enough residuals for GARCH")
    gfit = arch_model(resids, vol="Garch", p=1, q=1, rescale=False).fit(disp="off")
    shock = gfit.simulate(gfit.params, nobs=1, repetitions=MC_N).data.values.ravel()
    base  = float(model.predict(1, X=exo_row)[0]) if exo_row is not None else float(model.predict(1)[0])
    path  = base + shock
    stats = dict(Mean=path.mean(),
                 SD=path.std(),
                 P05=np.quantile(path, .05),
                 P95=np.quantile(path, .95))
    return path, stats

def _vs_plot(df, scen, colour, dates):
    png_path = os.path.join(tempfile.gettempdir(), f"{scen}_HYB.png")
    plt.figure(figsize=(10,5), dpi=150)
    plt.scatter(dates, df["Actual"], s=12, color="black", label="Actual")
    plt.plot(dates, df["Prediction"], lw=1, color=colour, label="Hybrid")
    plt.title(f"{scen} Hybrid – Actual vs Forecast", font="Times New Roman")
    plt.xlabel(""); plt.ylabel("Return"); plt.legend()
    plt.tight_layout(); plt.savefig(png_path); plt.close()
    return png_path

def walk_forward_hybrid(y, Xbest, scenario, base_fit, xgb_model):
    if Xbest.size == 0:
        Xbest = np.empty((len(y),0))
    def _one_split(i):
        y_tr = y[i-WIN_WF:i]
        X_tr = Xbest[i-WIN_WF:i]
        if   scenario=="lag0":  X_te = None if X_tr.size==0 else Xbest[i:i+1]
        elif scenario=="lag12": X_te = None if X_tr.size==0 or i+1-LAG_12<=0 else Xbest[i+1-LAG_12:i+2-LAG_12]
        else:                   X_te = None if X_tr.size==0 else _forecast_exo_future(Xbest, X_tr.shape[0])
        use_x = X_te is not None and X_tr.size
        mdl = _auto_sarima(y_tr, X_tr if use_x else None)
        arima_fc = mdl.predict(1, X=X_te)[0] if use_x else mdl.predict(1)[0]
        hyb_fc = arima_fc + _predict_xgb(xgb_model, X_te)
        act = y[i]
        return dict(SplitEnd=i-1,
                    Prediction=hyb_fc,
                    Actual=act,
                    RMSE=abs(hyb_fc-act),
                    Test_Hit=int(np.sign(hyb_fc)==np.sign(act)))
    iter_idx = range(WIN_WF, len(y)-1, STEP_WF)
    logs = Parallel(n_jobs=N_CPU, backend="loky")(delayed(_one_split)(i) for i in tqdm(iter_idx, desc=f"WF-{scenario}", unit="split", dynamic_ncols=True))
    return pd.DataFrame(logs)

def main():
    # ─── LOAD & PREP DATA ────────────────────────────────────────────────────────
    raw = pd.read_excel(IN_FILE, header=0, skiprows=2)
    raw = raw.iloc[::-1].reset_index(drop=True)
    date_vec = pd.to_datetime("20" + raw.iloc[:,0].str[:2] + "-" + raw.iloc[:,0].str[3:5] + "-01")
    ret = _to_num(raw.iloc[:, 7])
    base_feats = ["Lag1", "Lag2", "Lag3"]
    other_columns = [5,6,8,9,11,12,13,15,16,17,18]
    exo_cols = other_columns  # for column selection
    exo_df = raw.iloc[:, exo_cols].apply(_to_num)
    for lag in base_feats:
        exo_df[lag] = ret.shift(int(lag[-1]))
    # Reorder columns to force lags first
    exo_df = exo_df[base_feats + [c for c in exo_df.columns if c not in base_feats]]

    assert exo_df.notna().any().all(), "Some X columns are all-NaN"
    print("X-matrix shape:", exo_df.shape)

    keep = ret.notna()  # keep rows where the target exists
    exo_df = exo_df.fillna(method="ffill")  # or .interpolate(), or any imputation you prefer

    ret = ret[keep].to_numpy()
    date_vec = date_vec[keep]
    exo = exo_df[keep].to_numpy()
    idx_t = exo.shape[0]

    print(f"{exo.shape[1]} exogenous columns after cleaning")
    if exo.shape[1] == 0:
        raise ValueError("No usable exogenous variables – check the data prep")

    # ─── GRID SEARCH → BEST SARIMA PER SCENARIO ──────────────────────────────────
    print("\nGrid search → Lag-0 …")
    df0  = _grid_search(ret, exo, "lag0")
    print("Grid search → Lag-12 …")
    df12 = _grid_search(ret, exo, "lag12")
    print("Grid search → ARIMAX …")
    dfAR = _grid_search(ret, exo, "arimax")

    sel0  = _parse_cols(df0.iloc[0]["ExoSubset"])
    sel12 = _parse_cols(df12.iloc[0]["ExoSubset"])
    selAR = _parse_cols(dfAR.iloc[0]["ExoSubset"])

    print("Lag-0 exo subset:", sel0)
    print("Lag-12 exo subset:", sel12)
    print("ARIMAX exo subset:", selAR)

    X0  = exo[:, sel0 ] if sel0  else np.empty((idx_t,0))
    X12 = exo[:, sel12] if sel12 else np.empty((idx_t,0))
    XAR = exo[:, selAR] if selAR else np.empty((idx_t,0))

    print("X0 shape:", X0.shape, "Sample:", X0[:2])
    print("X12 shape:", X12.shape, "Sample:", X12[:2])
    print("XAR shape:", XAR.shape, "Sample:", XAR[:2])

    fit0  = _auto_sarima(ret, X0  if X0.size  else None)
    fit12 = _auto_sarima(ret, X12 if X12.size else None)
    fitAR = _auto_sarima(ret, XAR if XAR.size else None)

    # ─── STAGE-2 XGBOOST ON ONE-STEP RESIDUALS ───────────────────────────────────
    res0  = ret - fit0.predict_in_sample(X=X0 if X0.size else None)
    res12 = ret - fit12.predict_in_sample(X=X12 if X12.size else None)
    resAR = ret - fitAR.predict_in_sample(X=XAR if XAR.size else None)

    xgb0  = _best_xgb(X0 , res0)  if X0.size  else None
    xgb12 = _best_xgb(X12, res12) if X12.size else None
    xgbAR = _best_xgb(XAR, resAR) if XAR.size else None

    # ─── LIVE HYBRID FORECASTS ───────────────────────────────────────────────────
    exo_live0  = X0[-1:]  if X0.size  else None
    exo_live12 = X12[-LAG_12:-LAG_12+1] if X12.size and idx_t-LAG_12>=0 else None
    exo_liveAR = _forecast_exo_future(XAR, idx_t) if XAR.size else None

    live_arima0  = float(fit0.predict(1, X=exo_live0 )[0] if X0.size  else fit0.predict(1)[0])
    live_arima12 = float(fit12.predict(1, X=exo_live12)[0] if X12.size else fit12.predict(1)[0])
    live_arimaAR = float(fitAR.predict(1, X=exo_liveAR)[0] if XAR.size else fitAR.predict(1)[0])

    live_hyb0  = live_arima0  + _predict_xgb(xgb0 , exo_live0 )
    live_hyb12 = live_arima12 + _predict_xgb(xgb12, exo_live12)
    live_hybAR = live_arimaAR + _predict_xgb(xgbAR , exo_liveAR)

    # ─── WALK-FORWARD HYBRID BACK-TEST ───────────────────────────────────────────
    wf0  = walk_forward_hybrid(ret, X0 , "lag0",  fit0, xgb0 )
    wf12 = walk_forward_hybrid(ret, X12, "lag12", fit12, xgb12)
    wfAR = walk_forward_hybrid(ret, XAR, "arimax",fitAR, xgbAR)

    # ─── BUILD EXCEL WORKBOOK ────────────────────────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    def _sheet(df, name):
        ws = wb.create_sheet(name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        for row in ws.iter_rows():
            for c in row: c.font = Font(name="Times New Roman", size=12)

    _sheet(df0 , "Lag-0 Subsets")
    _sheet(df12, "Lag-12 Subsets")
    _sheet(dfAR, "ARIMAX Subsets")
    _sheet(wf0 , "Lag-0_WF")
    _sheet(wf12, "Lag-12_WF")
    _sheet(wfAR, "ARIMAX_WF")
    _sheet(pd.DataFrame({
            "Model":["Lag-0","Lag-12","ARIMAX"],
            "HitRate":[_hit_rate(wf0), _hit_rate(wf12), _hit_rate(wfAR)]
          }), "WF_HitRates")
    _sheet(pd.DataFrame({
            "Model":["Lag-0","Lag-12","ARIMAX"],
            "HybridForecast":[live_hyb0, live_hyb12, live_hybAR]
          }), "Live_Forecast")

    # ─── PNG DIAGNOSTICS ─────────────────────────────────────────────────────────
    png0  = _vs_plot(wf0 ,"Lag-0" , "#1f77b4", date_vec[wf0 ["SplitEnd"]+1])
    png12 = _vs_plot(wf12,"Lag-12", "#2ca02c", date_vec[wf12["SplitEnd"]+1])
    pngAR = _vs_plot(wfAR,"ARIMAX", "#d62728", date_vec[wfAR["SplitEnd"]+1])

    def _embed(sheet, path):
        if not os.path.exists(path): return
        ws = wb.create_sheet(sheet)
        img = XLImage(path); img.anchor = "B2"; ws.add_image(img)
        for row in ws.iter_rows():
            for c in row: c.font = Font(name="Times New Roman", size=12)

    _embed("Lag-0_VS" , png0 )
    _embed("Lag-12_VS", png12)
    _embed("ARIMAX_VS", pngAR)

    # ─── SAVE ────────────────────────────────────────────────────────────────────
    wb.save(OUT_FILE)
    print(f"\nWorkbook saved → {OUT_FILE}")

if __name__ == "__main__":
    pr = cProfile.Profile(); pr.enable(); t0 = time.time()
    main()
    pr.disable();       elapsed = time.time() - t0
    s = io.StringIO(); pstats.Stats(pr, stream=s).sort_stats("cumtime").print_stats(30)
    print(s.getvalue()); print(f"Total elapsed: {elapsed:.1f} s")