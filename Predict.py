#!/usr/bin/env python3
"""
Daily prediction script for XGBoost model.
Run this script daily to get next day's prediction.
"""

import joblib
import json
import xgboost as xgb
import pandas as pd
import numpy as np
from pathlib import Path
import talib as ta

# ─── CONSTANTS ──────────────────────────────────────────────────────────────
MODEL_FILE = "xgb_final_model.pkl"        # Saved model from training
CONFIG_FILE = "xgb_config.json"           # Saved config from training
LATEST_DATA = "/Users/shanechiu/Downloads/Index_DT1.xlsx"  # Update this daily
OUTPUT_FILE = "/Users/shanechiu/Desktop/prediction_log.xlsx"  # Where to save predictions

def _tech_indicators(high: pd.Series,
                     low: pd.Series,
                     close: pd.Series,
                     volume: pd.Series,
                     lb: int) -> pd.DataFrame:
    """Return a feature block for one look-back (lb)."""
    macd, macd_signal, _ = ta.MACD(close, 12, 26, 9)

    return pd.DataFrame({
        # ── price trend ───────────────────────────
        f"SMA_{lb}": ta.SMA(close, lb),
        f"EMA_{lb}": ta.EMA(close, lb),
        f"ROC_{lb}": ta.ROC(close, lb),

        # ── momentum / oscillators ────────────────
        f"RSI_{lb}":     ta.RSI(close, lb),
        f"ADX_{lb}":     ta.ADX(high, low, close, lb),
        f"MACD_hist":    macd - macd_signal,

        # ── volatility context ────────────────────
        f"ATR_{lb}":     ta.ATR(high, low, close, lb) / close,
        f"BBW_{lb}":     (lambda u,m,l: (u - l) / m)(*ta.BBANDS(close, lb)),

        # ── volume pressure ───────────────────────
        "OBV":           ta.OBV(close, volume),

        # ── autoregressive memory ────────────────
        "Ret_lag1":      close.pct_change().mul(100).shift(1),
    })

def main():
    # Load model and config
    model = joblib.load(MODEL_FILE)
    with open(CONFIG_FILE, "r") as f:
        config = json.load(f)
    
    # Load latest data (update this file daily)
    df = pd.read_excel(LATEST_DATA)
    
    # Compute features
    feats = _tech_indicators(
        df['High'], df['Low'], df['Close'], df['Volume'], 
        lb=config["BestLB"]
    ).dropna()
    
    # Get the last row for prediction
    X_new = feats.tail(1).to_numpy()
    dtest = xgb.DMatrix(X_new, feature_names=config["FeatureNames"])
    
    # Predict
    pred_ret = float(model.predict(dtest)[0])
    print(f"\nPredicted return: {pred_ret:.4f} %")
    
    # Position logic
    side = 1 if pred_ret >= 0 else -1
    print(f"Suggested Position Side: {'Long' if side == 1 else 'Short'}")
    
    # Calculate position size based on volatility
    ret_std = df['Close'].pct_change().mul(100).std()
    raw_size = abs(pred_ret) / ret_std if ret_std > 0 else 0.0
    size = min(raw_size, 2.0)  # Max 2x leverage
    print(f"Suggested Position Size: {size:.2f}x")

    # Save prediction to Excel
    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from datetime import datetime

    # Create prediction log
    pred_df = pd.DataFrame({
        'Date': [datetime.now().strftime('%Y-%m-%d')],
        'Predicted_Return': [pred_ret],
        'Position_Side': ['Long' if side == 1 else 'Short'],
        'Position_Size': [size]
    })

    # Try to load existing log or create new one
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Predictions"
        
        # Add headers
        for col, header in enumerate(pred_df.columns, 1):
            ws.cell(row=1, column=col, value=header)
            ws.cell(row=1, column=col).font = Font(bold=True)
        
        # Add new prediction
        for row in dataframe_to_rows(pred_df, index=False, header=False):
            ws.append(row)
        
        wb.save(OUTPUT_FILE)
        print(f"\nPrediction saved to {OUTPUT_FILE}")
    except Exception as e:
        print(f"Error saving prediction: {e}")

if __name__ == "__main__":
    main() 