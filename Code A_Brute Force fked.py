#!/opt/homebrew/Caskroom/miniconda/base/envs/py312env/bin/python
###############################################################################
#  MONTHLY  SARIMA–GARCH  PIPELINE  – PYTHON EDITION (Actual = Dots)
#
#  • 3 scenarios: Lag-0, Lag-12, ARIMAX
#  • Rolling-origin grid search (window = 241, step = 5) ranked by
#    RMSE → MASE → HitRate
#  • Parallel subset search (joblib)
#  • Monte-Carlo GARCH fan (3 000 paths) for live-risk view
#  • Walk-forward back-test (window = 241, step = 1) over **every** split
#  • **PNG sheets**: Actual (black dots) vs forecast (coloured lines) with legend
#  • Outputs a Times New Roman, 12-pt Excel workbook
#
#  ── REQUIRED LIBRARIES ──────────────────────────────────────────────────────
#  pip install pandas numpy openpyxl xlsxwriter matplotlib pillow joblib
#              pmdarima statsmodels arch tqdm
###############################################################################

from __future__ import annotations
import itertools, math, os, string, warnings, tempfile, datetime as dt
import numpy as np, pandas as pd
import matplotlib.pyplot as plt
from joblib import Parallel, delayed, cpu_count
from tqdm import tqdm
import pmdarima as pm
from statsmodels.tsa.statespace.sarimax import SARIMAX
from arch import arch_model
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font
import cProfile, pstats, io, time
from statsforecast.models import AutoARIMA

warnings.filterwarnings("ignore")
os.environ["PYTHONWARNINGS"] = "ignore"

# ─────────────── PARAMS (EDIT HERE) ──────────────────────────────────────────
IN_FILE      = "/Users/shanechiu/Downloads/Index_MT2.xlsx"
OUT_FILE     = "/Users/shanechiu/Desktop/2023_01M_ALL_CodeA.xlsx"

WIN_SIZE     = 241       # grid-search rolling window
STEP_SIZE_RO = 5         # grid-search step
WIN_WF       = 241       # walk-forward training window
STEP_WF      = 1         # walk-forward step
MC_N         = 3_000
MC_SEED      = 123
LAG_12       = 12        # months for "same month last year"
N_CPU        = cpu_count()
N_JOBS       = N_CPU
# ─────────────────────────────────────────────────────────────────────────────

# ─────────────── HELPERS ─────────────────────────────────────────────────────
def _parse_cols(s: str) -> list[int]:
    return [] if pd.isna(s) or s in ("", "None") else list(map(int, s.split(",")))

def _to_num(col: pd.Series) -> pd.Series:
    return pd.to_numeric(col.astype(str).str.replace("%", ""), errors="coerce")

def _hit_rate(df: pd.DataFrame) -> float:
    if df.empty or "Test_Hit" not in df.columns:
        return np.nan
    return df["Test_Hit"].mean()

def _auto_sarima(y: np.ndarray, x: np.ndarray | None):
    """
    Wrapper around Nixtla's lightning-fast AutoARIMA (numba-compiled).
    Keeps the same signature the rest of the script expects.
    """
    # StatsForecast's AutoARIMA handles exogenous vars via X=
    model = AutoARIMA(season_length=12)          # same seasonal spec
    if x is not None and x.size:
        model = model.fit(y=y, X=x)              # with exogenous
    else:
        model = model.fit(y=y)                   # univariate
    return model

def _forecast_exo_future(x: np.ndarray, idx_t: int) -> np.ndarray | None:
    """Forecast 1-step ahead for each exogenous column (ARIMA)"""
    if x.size == 0: return None
    fut = []
    for j in range(x.shape[1]):
        try:
            fit = pm.auto_arima(
                x[:idx_t, j], seasonal=True, stepwise=True,
                suppress_warnings=True, error_action="ignore"
            )
            fut.append(float(fit.predict(1)))
        except Exception:
            fut.append(np.nan)
    return np.array(fut).reshape(1, -1)

def _roll_origin_metrics(y: np.ndarray,
                         x: np.ndarray,
                         scenario: str,
                         window_size: int = WIN_SIZE,
                         step_size: int = STEP_SIZE_RO,
                         lag_val: int = LAG_12) -> tuple[float, float, float]:
    n = len(y)
    if n <= window_size: return (np.nan, np.nan, np.nan)
    se = ae = naive_ae = hits = tot = 0
    for i in range(window_size, n - 1, step_size):
        y_tr = y[i - window_size:i]
        X_tr = x[i - window_size:i]
        if   scenario == "lag0":   X_te = None if X_tr.size == 0 else X_tr[-1:]
        elif scenario == "lag12":  X_te = None if X_tr.size == 0 or i + 1 - lag_val <= 0 else X_tr[-lag_val:-lag_val+1]
        else:  # arimax
            X_te = None if X_tr.size == 0 else _forecast_exo_future(X_tr, X_tr.shape[0])
        use_x = X_te is not None and X_tr.size != 0
        try:
            model = _auto_sarima(y_tr, X_tr if use_x else None)
            fc = model.predict(1, X=X_te)[0] if use_x else model.predict(1)[0]
        except Exception:
            continue
        act = y[i]
        if not math.isnan(fc) and not math.isnan(act):
            se       += (fc - act) ** 2
            ae       += abs(fc - act)
            naive_ae += abs(y_tr[-1] - act)
            hits     += int(np.sign(fc) == np.sign(act))
            tot      += 1
    if tot == 0: return (np.nan, np.nan, np.nan)
    return round(math.sqrt(se / tot), 4), round(ae / naive_ae, 4), round(hits / tot, 4)

def _grid_search(y: np.ndarray,
                 x: np.ndarray,
                 scenario: str) -> pd.DataFrame:
    n_exo = x.shape[1]
    combos = range(2 ** n_exo)  # bit mask
    def _eval(mask: int) -> tuple[str, float, float, float]:
        cols = [i for i in range(n_exo) if mask & (1 << i)]
        x_mat = x[:, cols] if cols else np.empty((len(y), 0))
        rmse, mase, hit = _roll_origin_metrics(y, x_mat, scenario)
        subset = ",".join(map(str, cols)) if cols else "None"
        return subset, rmse, mase, hit
    results = Parallel(n_jobs=N_JOBS, backend="loky")(
        delayed(_eval)(m) for m in tqdm(combos, desc=f"Grid-{scenario} (2^{n_exo})")
    )
    df = pd.DataFrame(results, columns=["ExoSubset", "RMSE", "MASE", "HitRate"])
    df = df.sort_values(["RMSE", "MASE", "HitRate"], ascending=[True, True, False])
    df.insert(0, "Rank", range(1, len(df) + 1))
    return df

def _walk_forward(y: np.ndarray,
                  x: np.ndarray,
                  scenario: str) -> pd.DataFrame:
    n = len(y)
    def _one_split(i):
        y_tr = y[i - WIN_WF:i]
        X_tr = x[i - WIN_WF:i]
        if   scenario == "lag0":   X_te = None if X_tr.size == 0 else X_tr[-1:]
        elif scenario == "lag12":  X_te = None if X_tr.size == 0 or i + 1 - LAG_12 <= 0 else X_tr[-LAG_12:-LAG_12+1]
        else:  # arimax
            X_te = None if X_tr.size == 0 else _forecast_exo_future(X_tr, X_tr.shape[0])
        use_x = X_te is not None and X_tr.size != 0
        try:
            model = _auto_sarima(y_tr, X_tr if use_x else None)
            fc = model.predict(1, X=X_te)[0] if use_x else model.predict(1)[0]
        except Exception:
            return None
        act = y[i]
        return dict(
            SplitEnd=i-1,
            Prediction=fc,
            Actual=act,
            RMSE=abs(fc - act),
            Test_Hit=int(np.sign(fc) == np.sign(act))
        )
    idx = range(WIN_WF, n - 1, STEP_WF)
    logs = Parallel(n_jobs=N_CPU, backend="loky")(
        delayed(_one_split)(i) for i in tqdm(idx, desc=f"WF-{scenario}", unit="split", dynamic_ncols=True)
    )
    logs = [log for log in logs if log is not None]
    return pd.DataFrame(logs)

def _monte_carlo(model: pm.ARIMA | SARIMAX,
                 exo_row: np.ndarray | None) -> tuple[np.ndarray, dict[str, float]]:
    np.random.seed(MC_SEED)
    resids = model.arima_res_.resid
    if len(resids) < 20:
        raise ValueError("Not enough residuals for GARCH")
    garch = arch_model(resids, vol="Garch", p=1, q=1, rescale=False)
    gfit  = garch.fit(disp="off")
    shock = gfit.simulate(gfit.params, nobs=1, repetitions=MC_N).data.values.ravel()
    base  = float(model.predict(1, X=exo_row)[0]) if exo_row is not None else float(model.predict(1)[0])
    path  = base + shock
    stats = dict(Mean=path.mean(),
                 SD=path.std(),
                 P05=np.quantile(path, 0.05),
                 P95=np.quantile(path, 0.95))
    return path, stats

def _vs_plot(df: pd.DataFrame, scen: str, colour: str, dates: pd.Series) -> str:
    temp_png = os.path.join(tempfile.gettempdir(), f"{scen}_VS.png")
    plt.figure(figsize=(10, 5), dpi=150)
    plt.scatter(dates, df["Actual"], s=12, label="Actual", color="black")
    plt.plot   (dates, df["Prediction"], lw=1, label=scen, color=colour)
    plt.title(f"{scen} – Actual vs Forecast", font="Times New Roman")
    plt.xlabel(""); plt.ylabel("Return"); plt.legend()
    plt.tight_layout(); plt.savefig(temp_png); plt.close()
    return temp_png

def main():
    # ─────────────── LOAD & PREP DATA ────────────────────────────────────────────
    raw = pd.read_excel(IN_FILE, header=0, skiprows=2)
    raw = raw.iloc[::-1].reset_index(drop=True)        # chronological
    date_vec = pd.to_datetime("20" + raw.iloc[:, 0].str[:2] + "-" +
                              raw.iloc[:, 0].str[3:5] + "-01")
    ret = _to_num(raw.iloc[:, 7])
    exo_cols = [5, 6, 8, 9, 11, 12, 13, 15, 16, 17, 18]  # 0-based indexing
    exo_df = raw.iloc[:, exo_cols].apply(_to_num)
    exo_df["Lag1"] = ret.shift(1)
    exo_df["Lag2"] = ret.shift(2)
    exo_df["Lag3"] = ret.shift(3)
    keep = ret.notna() & exo_df.notna().all(axis=1)
    ret = ret[keep].to_numpy()
    date_vec = date_vec[keep]
    exo = exo_df[keep].to_numpy()
    idx_t = exo.shape[0]

    # ─────────────── GRID SEARCH ─────────────────────────────────────────────────
    print("\nGrid search → Lag-0 …")
    df_lag0  = _grid_search(ret, exo, "lag0")
    print("Grid search → Lag-12 …")
    df_lag12 = _grid_search(ret, exo, "lag12")
    print("Grid search → ARIMAX …")
    df_ar    = _grid_search(ret, exo, "arimax")

    # ────── BEST FITS ────────────────────────────────────────────────────────────
    best0_idx  = _parse_cols(df_lag0.iloc[0]["ExoSubset"])
    best12_idx = _parse_cols(df_lag12.iloc[0]["ExoSubset"])
    bestAR_idx = _parse_cols(df_ar.iloc[0]["ExoSubset"])

    Xbest0  = exo[:, best0_idx]  if best0_idx  else np.empty((idx_t, 0))
    Xbest12 = exo[:, best12_idx] if best12_idx else np.empty((idx_t, 0))
    XbestAR = exo[:, bestAR_idx] if bestAR_idx else np.empty((idx_t, 0))

    fit0  = _auto_sarima(ret, Xbest0  if Xbest0.size  else None)
    fit12 = _auto_sarima(ret, Xbest12 if Xbest12.size else None)
    fitAR = _auto_sarima(ret, XbestAR if XbestAR.size else None)

    exo_live0      = Xbest0[-1:]  if Xbest0.size  else None
    exo_live12     = Xbest12[-LAG_12:-LAG_12+1] if Xbest12.size and idx_t - LAG_12 >= 0 else None
    exo_future_AR  = _forecast_exo_future(XbestAR, idx_t) if XbestAR.size else None

    # ─────────────── EXCEL WORKBOOK ──────────────────────────────────────────────
    wb = Workbook(write_only=False)
    ws0  = wb.create_sheet("Lag-0 Subsets");   ws12 = wb.create_sheet("Lag-12 Subsets")
    wsAR = wb.create_sheet("ARIMAX Subsets")
    for ws in (ws0, ws12, wsAR):
        ws.sheet_view.showGridLines = False

    def _dump(df: pd.DataFrame, sheet):
        for r in dataframe_to_rows(df, index=False, header=True):
            sheet.append(r)

    _dump(df_lag0,  ws0);  _dump(df_lag12, ws12);  _dump(df_ar,  wsAR)

    # Best subset + Monte-Carlo
    def _write_best(name, df_sub, fit, exo_row):
        ws = wb.create_sheet(f"{name} Best")
        _dump(df_sub.iloc[[0]], ws)
        try:
            path, stats = _monte_carlo(fit, exo_row)
            ws_mc = wb.create_sheet(f"{name}_MC")
            mc_df = pd.DataFrame({"Sim": np.arange(1, MC_N+1), "Path": path})
            _dump(mc_df, ws_mc)
            stats_df = pd.DataFrame(stats, index=[0]).T.reset_index()
            stats_df.columns = ["Metric", "Value"]
            _dump(stats_df, ws_mc)
        except Exception as e:
            ws.append(["Monte-Carlo error", str(e)])

    _write_best("Lag-0",   df_lag0,  fit0,  exo_live0)
    _write_best("Lag-12",  df_lag12, fit12, exo_live12)
    _write_best("ARIMAX",  df_ar,    fitAR, exo_future_AR)

    # Walk-forward logs
    wf0  = _walk_forward(ret, Xbest0,  "lag0")
    wf12 = _walk_forward(ret, Xbest12, "lag12")
    wfAR = _walk_forward(ret, XbestAR, "arimax")

    for name, df in [("Lag-0_WF", wf0), ("Lag-12_WF", wf12), ("ARIMAX_WF", wfAR)]:
        ws = wb.create_sheet(name); _dump(df, ws)

    hit_df = pd.DataFrame({
        "Model": ["Lag-0", "Lag-12", "ARIMAX"],
        "HitRate": [_hit_rate(wf0), _hit_rate(wf12), _hit_rate(wfAR)]
    })
    ws_hit = wb.create_sheet("WF_HitRates"); _dump(hit_df, ws_hit)

    pred = fit0.predict(1, X=exo_live0) if exo_live0 is not None else fit0.predict(1)
    if hasattr(pred, "iloc"):
        live_fc0 = float(pred.iloc[0])
    elif isinstance(pred, dict):
        for key in ["mean", "fcst", "forecast"]:
            if key in pred:
                val = pred[key]
                if isinstance(val, (np.ndarray, list, pd.Series)):
                    live_fc0 = float(val[0])
                else:
                    live_fc0 = float(val)
                break
        else:
            raise ValueError(f"Unknown dict format for prediction: {pred}")
    else:
        try:
            live_fc0 = float(pred[0])
        except (TypeError, IndexError, KeyError):
            live_fc0 = float(pred)

    pred = fit12.predict(1, X=exo_live12) if exo_live12 is not None else fit12.predict(1)
    if hasattr(pred, "iloc"):
        live_fc12 = float(pred.iloc[0])
    elif isinstance(pred, dict):
        for key in ["mean", "fcst", "forecast"]:
            if key in pred:
                val = pred[key]
                if isinstance(val, (np.ndarray, list, pd.Series)):
                    live_fc12 = float(val[0])
                else:
                    live_fc12 = float(val)
                break
        else:
            raise ValueError(f"Unknown dict format for prediction: {pred}")
    else:
        try:
            live_fc12 = float(pred[0])
        except (TypeError, IndexError, KeyError):
            live_fc12 = float(pred)

    pred = fitAR.predict(1, X=exo_future_AR) if exo_future_AR is not None else fitAR.predict(1)
    if hasattr(pred, "iloc"):
        live_fcAR = float(pred.iloc[0])
    elif isinstance(pred, dict):
        for key in ["mean", "fcst", "forecast"]:
            if key in pred:
                val = pred[key]
                if isinstance(val, (np.ndarray, list, pd.Series)):
                    live_fcAR = float(val[0])
                else:
                    live_fcAR = float(val)
                break
        else:
            raise ValueError(f"Unknown dict format for prediction: {pred}")
    else:
        try:
            live_fcAR = float(pred[0])
        except (TypeError, IndexError, KeyError):
            live_fcAR = float(pred)

    live_df = pd.DataFrame({"Model": ["Lag-0","Lag-12","ARIMAX"],
                            "Forecast": [live_fc0, live_fc12, live_fcAR]})
    ws_live = wb.create_sheet("Live_Forecast"); _dump(live_df, ws_live)

    # ─────────────── PNG DIAGNOSTICS ─────────────────────────────────────────────
    png0  = _vs_plot(wf0,  "Lag-0",  "#1f77b4", date_vec[wf0["SplitEnd"]+1])
    png12 = _vs_plot(wf12, "Lag-12", "#2ca02c", date_vec[wf12["SplitEnd"]+1])
    pngAR = _vs_plot(wfAR, "ARIMAX", "#d62728", date_vec[wfAR["SplitEnd"]+1])

    def _insert_png(sheet_name: str, png_path: str):
        ws = wb.create_sheet(sheet_name)
        img = XLImage(png_path)
        img.anchor = "B2"
        ws.add_image(img)

    _insert_png("Lag-0_VS",  png0)
    _insert_png("Lag-12_VS", png12)
    _insert_png("ARIMAX_VS", pngAR)

    # ─────────────── FORMATTING ──────────────────────────────────────────────────
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.font = Font(name="Times New Roman", size=12)

    # ─────────────── SAVE ────────────────────────────────────────────────────────
    wb.remove(wb["Sheet"])  # default sheet
    wb.save(OUT_FILE)
    print(f"\nWorkbook saved → {OUT_FILE}")

if __name__ == "__main__":
    pr = cProfile.Profile(); pr.enable(); t0 = time.time()
    main()
    pr.disable();   elapsed = time.time() - t0
    s = io.StringIO(); pstats.Stats(pr, stream=s).sort_stats("cumtime").print_stats(30)
    print(s.getvalue()); print(f"Elapsed: {elapsed:.1f} s")