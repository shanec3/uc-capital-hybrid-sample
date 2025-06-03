# -*- coding: utf-8 -*-
"""
DAILY  TECHNICAL-XGBOOST  PIPELINE – PYTHON EDITION (regression version)

* Predict next-day % return (objective = reg:squarederror)
* Rolling-origin walk-forward with Optuna hyper-param search
* Position side = sign(predicted return)
* Position size = |pred_ret| / rolling σ, clipped to ≤ 2×
* Trade-level stop widened to –5 %
* Grid-search over training-window length & indicator look-back, ranked by **Sharpe**
* Outputs: walk-forward log, KPI sheet (CAGR, Sharpe, MDD, total PnL), SHAP importances, yearly PNGs

Assumptions
-----------
* Daily data; date column formatted "DD/MM/YY".
* Worksheet header row starts in row 3 (skiprows=2) like the original workbook.
* Price is in the 3rd column (index 2).

Dependencies
------------
$ pip install pandas numpy openpyxl xlsxwriter matplotlib pillow optuna shap tqdm joblib talib-binary xgboost
"""

from __future__ import annotations
import os, math, tempfile, warnings, time, random, datetime as dt
from pathlib import Path

import numpy as np, pandas as pd
import matplotlib.pyplot as plt
from joblib import cpu_count, Parallel, delayed
from tqdm import tqdm
import talib as ta
import xgboost as xgb
from sklearn.model_selection import TimeSeriesSplit
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import optuna
import shap
import logging

# ─── CONSTANTS ──────────────────────────────────────────────────────────────
SEED       = 123
IN_FILE    = "/Users/shanechiu/Downloads/Index_DT1.xlsx"   # change me
OUT_FILE   = "/Users/shanechiu/Desktop/2025_0101_XGB.xlsx"  # change me

STEP_WF    = 1                    # 1 bar forward each split
EARLY_STOP = 50                   # XGB early-stopping rounds
ANNUAL_F   = math.sqrt(252)       # daily ⇒ annual

# candidate training-window lengths (number of bars)
CAND_TRAIN = [252, 504, 756, 1260]        # 1, 2, 3, 5 years
# indicator look-back periods (in bars)
CANDIDATE_LB = [4, 8, 13, 26, 52]

# leverage & risk
MAX_LEV    = 2.0                  # ≤ 2× notional
STOP_PCT   = -5.0                 # hard stop (–5 % trade-level cap)

random.seed(SEED)
np.random.seed(SEED)
warnings.filterwarnings("ignore")
os.environ["PYTHONWARNINGS"] = "ignore"

optuna.logging.set_verbosity(optuna.logging.CRITICAL)
logging.getLogger("optuna").setLevel(logging.CRITICAL)
logging.getLogger("optuna").propagate = False
optuna.logging.disable_default_handler()   # silence "Trial … finished"

# ─── HELPERS ────────────────────────────────────────────────────────────────

def _to_num(col: pd.Series) -> pd.Series:
    """Sanitise numbers that may contain commas, %, etc."""
    return pd.to_numeric(col.astype(str).str.replace(r"[^\d\.\-+]", "", regex=True), errors="coerce")


def _parse_date(s: str):
    """Parse DD/MM/YY → datetime, tolerate quotes/spaces."""
    s = str(s).strip().strip("'").strip('"')
    for fmt in ("%y/%m/%d", "%d/%m/%y"):          # year-first first
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return pd.NaT


def _tech_indicators(high: pd.Series,
                     low: pd.Series,
                     close: pd.Series,
                     volume: pd.Series,
                     lb: int) -> pd.DataFrame:
    """Return a feature block for one look-back (lb)."""
    macd, macd_signal, _ = ta.MACD(close, 12, 26, 9)

    return pd.DataFrame({
        # ── price trend ───────────────────────────
        f"SMA_{lb}": ta.SMA(close, lb),
        f"EMA_{lb}": ta.EMA(close, lb),
        f"ROC_{lb}": ta.ROC(close, lb),

        # ── momentum / oscillators ────────────────
        f"RSI_{lb}":     ta.RSI(close, lb),
        f"ADX_{lb}":     ta.ADX(high, low, close, lb),
        f"MACD_hist":    macd - macd_signal,

        # ── volatility context ────────────────────
        f"ATR_{lb}":     ta.ATR(high, low, close, lb) / close,
        f"BBW_{lb}":     (lambda u,m,l: (u - l) / m)(*ta.BBANDS(close, lb)),

        # ── volume pressure ───────────────────────
        "OBV":           ta.OBV(close, volume),

        # ── autoregressive memory ────────────────
        "Ret_lag1":      close.pct_change().mul(100).shift(1),
    })


# ─── CORE WALK-FORWARD ──────────────────────────────────────────────────────

def _rolling_walk_forward(X: np.ndarray,
                           y: np.ndarray,
                           feature_names: list[str],
                           win_wf: int,
                           ret_vec: np.ndarray,
                           dates: pd.Series,
                           shap_prune_frac: float = 0.2) -> tuple[pd.DataFrame, pd.Series]:
    """Rolling-origin WF -> log DF & mean absolute SHAP importances."""
    idx_splits = list(range(win_wf, len(y)-1, STEP_WF))
    n_cpus     = cpu_count()
    n_feat     = X.shape[1]

    def wf_task(i: int):
        optuna.logging.disable_default_handler()   # needed again in each subprocess
        optuna.logging.set_verbosity(optuna.logging.WARNING)
        X_tr, y_tr = X[i-win_wf:i], y[i-win_wf:i]
        X_te, y_te = X[i:i+1],   y[i]

        # ── Optuna search (5-fold CV) ──
        def objective(trial):
            params = {
                'max_depth':        trial.suggest_int('max_depth', 2, 6),
                'eta':              trial.suggest_float('eta', 0.01, 0.3),
                'subsample':        trial.suggest_float('subsample', 0.5, 1.0),
                'colsample_bytree': trial.suggest_float('colsample_bytree', 0.5, 1.0),
                'objective':        'reg:squarederror',
                'tree_method':      'hist',
                'predictor':        'auto',
                'seed':             SEED,
                'verbosity':        0,
                'nthread':          n_cpus,
            }
            dtrain = xgb.DMatrix(X_tr, label=y_tr)
            cv_res = xgb.cv(params, dtrain, num_boost_round=500, nfold=5,
                            stratified=False, seed=SEED, early_stopping_rounds=EARLY_STOP,
                            verbose_eval=False)
            return cv_res['test-rmse-mean'].iloc[-1]

        study = optuna.create_study(direction='minimize')
        study.optimize(objective, n_trials=20, show_progress_bar=False)
        bst_params = study.best_params | {
            'objective': 'reg:squarederror',
            'tree_method': 'hist',
            'predictor': 'auto',
            'seed': SEED,
            'verbosity': 0,
            'nthread': n_cpus,
        }

        dtrain = xgb.DMatrix(X_tr, label=y_tr, feature_names=feature_names)
        dtest  = xgb.DMatrix(X_te, feature_names=feature_names)
        cv_res = xgb.cv(bst_params, dtrain, num_boost_round=1000, nfold=5,
                        stratified=False, seed=SEED, early_stopping_rounds=EARLY_STOP,
                        verbose_eval=False)
        n_round = len(cv_res)
        model   = xgb.train(bst_params, dtrain, num_boost_round=n_round)

        # prediction & trade metrics
        pred_ret = float(model.predict(dtest)[0])
        side     = 1 if pred_ret >= 0 else -1
        sigma    = ret_vec[i-win_wf:i].std(ddof=0)
        raw_size = abs(pred_ret) / sigma if sigma > 0 else 0.0
        size     = min(raw_size, MAX_LEV)
        pnl_raw  = side * size * y_te
        pnl      = max(pnl_raw, STOP_PCT)
        hit      = int(np.sign(pred_ret) == np.sign(y_te))

        # SHAP importance
        explainer   = shap.TreeExplainer(model, feature_perturbation="tree_path_dependent")
        shap_values = explainer.shap_values(X_tr)
        # ensure 2-D
        shap_arr = np.asarray(shap_values)
        if shap_arr.ndim == 1:
            shap_arr = shap_arr.reshape(-1, n_feat)
        mean_abs_shap = np.abs(shap_arr).mean(axis=0)

        err  = pred_ret - y_te
        se   = err ** 2
        return (
            {
                'DataDate': dates.iloc[i].strftime('%y%m%d'),
                'PredRet':  pred_ret,
                'Size':     size,
                'Side':     side,
                'Actual':   y_te,
                'Err':      err,
                'SE':       se,
                'PnL':      pnl,
                'Hit':      hit,
            },
            mean_abs_shap,
        )

    # run in parallel
    res = Parallel(n_jobs=n_cpus)(delayed(wf_task)(i) for i in tqdm(idx_splits, desc="Walk-Forward", unit="split"))

    if not res:
        return pd.DataFrame(), pd.Series(dtype=float)

    logs, shap_list = zip(*res)
    wf_df = pd.DataFrame(logs)
    shap_mat = np.vstack(shap_list)
    shap_mean = pd.Series(shap_mat.mean(axis=0), index=feature_names, name="MeanAbsSHAP")
    return wf_df, shap_mean


# ─── KPI & PLOT HELPERS ────────────────────────────────────────────────────

def _pnl_stats(pnl_pct: pd.Series):
    cum = pnl_pct.cumsum()
    peak = cum.cummax()
    dd   = peak - cum
    sharpe = pnl_pct.mean() / pnl_pct.std(ddof=0) * ANNUAL_F if pnl_pct.std(ddof=0) > 0 else np.nan
    return {
        'TotalPnL':    cum.iloc[-1],
        'MaxDrawdown': dd.max(),
        'Sharpe':      sharpe,
    }


def _save_yearly_plots(wf: pd.DataFrame, out_dir: Path) -> list[Path]:
    paths = []
    for yr, grp in wf.groupby(wf['Date'].str[:4]):
        if grp.empty:
            continue
        p = out_dir / f'TECH_XGB_VS_{yr}.png'
        plt.figure(figsize=(8,3), dpi=150)
        plt.scatter(grp['Date'], grp['Actual'], s=12, color='black')
        plt.plot(grp['Date'], grp['PredRet'], lw=1)
        plt.title(f'Actual vs Predicted – {yr}')
        plt.xlabel('Date'); plt.ylabel('Return %')
        plt.tight_layout(); plt.savefig(p); plt.close()
        paths.append(p)
    return paths


# ─── MAIN ───────────────────────────────────────────────────────────────────

def main():
    # ── LOAD ──
    raw = pd.read_excel(IN_FILE, header=None, skiprows=2)  # every row = data
    raw = raw.iloc[:, :6]  # Only keep the first 6 columns
    raw.columns = ["Date", "Open", "High", "Low", "Close", "Volume"]  # or whatever is correct
    raw = raw.drop(columns="Open")
    raw = raw.iloc[::-1].reset_index(drop=True)  # chronological ↑

    dates  = raw["Date"].apply(_parse_date)
    price  = _to_num(raw["Close"])
    ret    = price.pct_change().mul(100)           # %

    valid = dates.notna() & price.notna()
    dates = dates[valid].reset_index(drop=True)
    price = price[valid].reset_index(drop=True)
    ret = ret[valid].reset_index(drop=True)
    high = _to_num(raw["High"])[valid].reset_index(drop=True)
    low = _to_num(raw["Low"])[valid].reset_index(drop=True)
    volume = _to_num(raw["Volume"])[valid].reset_index(drop=True)

    best_sharpe = -np.inf
    best_params = None

    for win in CAND_TRAIN:
        for lb in CANDIDATE_LB:
            feats = _tech_indicators(high, low, price, volume, lb).dropna()
            df_all = pd.concat([dates.rename('Date'), price.rename('Close'), ret.rename('Ret'), feats], axis=1).dropna()
            if len(df_all) < win + 10:  # need enough samples
                continue
            X = df_all.iloc[:,3:].to_numpy()
            y = df_all['Ret'].values
            wf, _ = _rolling_walk_forward(X, y, feature_names=feats.columns.tolist(),
                                           win_wf=win, ret_vec=df_all['Ret'].values,
                                           dates=df_all['Date'])
            if wf.empty:
                continue
            sharpe = _pnl_stats(wf['PnL'])['Sharpe']
            if sharpe > best_sharpe:
                best_sharpe = sharpe
                best_params = (win, lb, wf, feats, df_all)
                # print(f"↪ new best Sharpe {sharpe:.2f} @ win={win}, lb={lb}")

    if best_params is None:
        print("No parameter set produced valid splits. Exiting.")
        return

    WIN_WF, BEST_LB, wf, feats, df_all = best_params

    # rerun one final time for SHAP on best params
    X = df_all.iloc[:,3:].to_numpy(); y = df_all['Ret'].values
    feats = _tech_indicators(high, low, price, volume, BEST_LB).dropna()
    wf, shap_imp = _rolling_walk_forward(X, y, feature_names=feats.columns.tolist(),
                                         win_wf=WIN_WF, ret_vec=df_all['Ret'].values,
                                         dates=df_all['Date'])

    wf.insert(1, 'Date', df_all['Date'].iloc[wf.index].dt.strftime('%Y-%m-%d'))
    # Move PredRet so it sits just before Actual in the WF log
    wf.insert(wf.columns.get_loc("Actual"), "PredRet", wf.pop("PredRet"))

    # Add RMSE
    rmse = np.sqrt(wf['SE'].mean())
    wf.attrs["RMSE"] = rmse  # handy if you want it later

    # KPI
    stats = _pnl_stats(wf['PnL'])
    start_cap = 1_000_000
    pnl_frac = wf['PnL']/100
    end_cap = start_cap * np.prod(1 + pnl_frac)
    years   = len(wf)/252
    cagr    = (end_cap/start_cap) ** (1/years) - 1
    total_pnl_pct = (end_cap/start_cap - 1) * 100

    # ── EXCEL ──
    wb = Workbook(); wb.remove(wb.active)

    def _sheet(df, name):
        ws = wb.create_sheet(name)
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        for row in ws.iter_rows():
            for c in row: c.font = Font(name="Times New Roman", size=12)

    _sheet(wf, "WF_Log")

    # write one extra row at the top with RMSE
    ws_log = wb["WF_Log"]
    ws_log.insert_rows(1)
    ws_log["A1"] = "RMSE:"
    ws_log["B1"] = rmse
    ws_log["A1"].font = ws_log["B1"].font = Font(name="Times New Roman", size=12, bold=True)

    kpi = pd.DataFrame({
        'Sharpe':        [stats['Sharpe']],
        'CAGR':          [cagr],
        'TotalPnL(%)':   [total_pnl_pct],
        'MaxDrawdown%':  [stats['MaxDrawdown']],
        'Start_$':       [start_cap],
        'End_$':         [end_cap],
        'BestWin':       [WIN_WF],
        'BestLB':        [BEST_LB],
    })
    # Format Start_$ and End_$ as plain numbers, not scientific
    kpi["Start_$"] = kpi["Start_$"].map("{:,.0f}".format)
    kpi["End_$"]   = kpi["End_$"].map("{:,.0f}".format)
    _sheet(kpi, "Summary_PnL")

    _sheet(shap_imp.reset_index(), "SHAP_Importances")

    # yearly plots
    tmp_dir = Path(tempfile.gettempdir())
    plot_paths = _save_yearly_plots(wf[['Date','Actual','PredRet']], tmp_dir)
    ws_img = wb.create_sheet('VS_Plots'); row = 2
    for p in plot_paths:
        img = XLImage(str(p)); img.anchor = f'B{row}'; ws_img.add_image(img); row += 25
    for r in ws_img.iter_rows():
        for c in r: c.font = Font(name="Times New Roman", size=12)

    wb.save(OUT_FILE)
    print(f"Workbook saved → {OUT_FILE}")


if __name__ == "__main__":
    t0 = time.time(); main(); print(f"Elapsed: {time.time()-t0:.1f}s")
