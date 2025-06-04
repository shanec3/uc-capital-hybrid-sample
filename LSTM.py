# -*- coding: utf-8 -*-
###############################################################################
#  DAILY  TECHNICAL-LSTM  PIPELINE – PYTHON EDITION  (regression version)
#
#  * Predict next-day % return with an LSTM neural network
#  * Rolling-origin walk-forward back-test with Optuna hyper-parameter search
#  * Position side  = sign(predicted return)
#  * Position size  = |pred_ret| / rolling σ, clipped to ≤ 2×
#  * Trade-level hard stop = –5 %
#  * Grid-search over training-window length & look-back sequence length,
#    ranked by **Sharpe ratio**
#  * Outputs: walk-forward log, KPI sheet (CAGR, Sharpe, MDD, total PnL),
#             yearly "Actual vs Pred" PNGs – all in a Times New Roman 12-pt
#             Excel workbook
#
#  Assumptions
#  -----------
#  * Daily data; date column formatted "DD/MM/YY".
#  * Worksheet header row starts in row 3 (skiprows=2) exactly like your file.
#  * Price is in the 3rd column (index 2).
#
#  Dependencies
#  ------------
#  $ pip install pandas numpy openpyxl xlsxwriter matplotlib pillow optuna tqdm \
#                joblib talib-binary scikit-learn tensorflow==2.* shap
###############################################################################

from __future__ import annotations
import os, math, tempfile, warnings, time, random, datetime as dt
from pathlib import Path
import numpy as np, pandas as pd
import matplotlib.pyplot as plt
from joblib import cpu_count
from tqdm import tqdm
import talib as ta
import optuna, logging
from sklearn.preprocessing import StandardScaler
from tensorflow import keras
from tensorflow.keras import layers, callbacks

# ── CONSTANTS ───────────────────────────────────────────────────────────────
SEED       = 123
IN_FILE    = "/Users/shanechiu/Downloads/Index_DT1.xlsx"    # ← change me
OUT_FILE   = "/Users/shanechiu/Desktop/2025_0101_LSTM.xlsx" # ← change me

STEP_WF    = 1                      # 1 bar forward each split
EARLY_STOP = 20                     # LSTM early-stopping patience
ANNUAL_F   = math.sqrt(252)         # daily → annualised

CAND_TRAIN = [252, 504, 756, 1260]  # 1, 2, 3, 5 years
CAND_SEQ   = [20, 40, 60]           # look-back sequence length (days)

MAX_LEV    = 2.0    # ≤ 2× notional
STOP_PCT   = -5.0   # trade-level hard stop (–5 %)

random.seed(SEED); np.random.seed(SEED)
os.environ["PYTHONHASHSEED"] = str(SEED)
warnings.filterwarnings("ignore")
optuna.logging.set_verbosity(optuna.logging.CRITICAL)
logging.getLogger("optuna").propagate = False

# ── HELPERS ────────────────────────────────────────────────────────────────
def _to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s.astype(str).str.replace(r"[^\d.\-+]", "", regex=True),
                         errors="coerce")

def _parse_date(txt: str):
    txt = str(txt).strip().strip("'").strip('"')
    for fmt in ("%y/%m/%d", "%d/%m/%y"):
        try:
            return dt.datetime.strptime(txt, fmt)
        except ValueError:
            continue
    return pd.NaT

def _tech_indicators(h, l, c, v, lb=14) -> pd.DataFrame:
    macd, macds, _ = ta.MACD(c, 12, 26, 9)
    return pd.DataFrame({
        f"SMA_{lb}": ta.SMA(c, lb),
        f"EMA_{lb}": ta.EMA(c, lb),
        f"RSI_{lb}": ta.RSI(c, lb),
        f"ATR_{lb}": ta.ATR(h, l, c, lb) / c,
        f"ROC_{lb}": ta.ROC(c, lb),
        "MACD_hist": macd - macds,
        "OBV":        ta.OBV(c, v),
        "Ret_lag1":   c.pct_change().mul(100).shift(1),
    })

def _create_seq(mat: np.ndarray, y: np.ndarray, seq: int):
    X, Y = [], []
    for i in range(seq, len(mat)):
        X.append(mat[i-seq:i]); Y.append(y[i])
    return np.asarray(X), np.asarray(Y)

def _pnl_stats(p: pd.Series):
    cum, peak = p.cumsum(), p.cumsum().cummax()
    dd = peak - cum
    sharpe = p.mean()/p.std(ddof=0)*ANNUAL_F if p.std(ddof=0)>0 else np.nan
    return {"TotalPnL": cum.iloc[-1], "MaxDrawdown": dd.max(), "Sharpe": sharpe}

def _plot_yearly(wf: pd.DataFrame, outdir: Path):
    paths=[]
    for yr, g in wf.groupby(wf['Date'].str[:4]):
        if g.empty: continue
        p=outdir/f"LSTM_VS_{yr}.png"
        plt.figure(figsize=(8,3),dpi=150)
        plt.scatter(g['Date'],g['Actual'],s=12,c='black')
        plt.plot(g['Date'],g['PredRet'],lw=1); plt.title(f"{yr}")
        plt.xlabel('Date');plt.ylabel('Return %');plt.tight_layout()
        plt.savefig(p);plt.close();paths.append(p)
    return paths

# ── WALK-FORWARD ────────────────────────────────────────────────────────────
def _walk(df: pd.DataFrame, win: int, seq: int)->pd.DataFrame:
    feats, targ = df.iloc[:,3:].values.astype(np.float32), df['Ret'].values.astype(np.float32)
    dates = df['Date'].reset_index(drop=True)
    idx = list(range(win+seq, len(df)-1, STEP_WF))
    scaler = StandardScaler(); logs=[]
    for i in tqdm(idx, desc=f"win={win} seq={seq}", unit="split"):
        X_raw = feats[i-win-seq:i]; y_raw=targ[i-win-seq:i]
        scaler.fit(X_raw); X_tr=scaler.transform(X_raw); X_te=scaler.transform(feats[i-seq:i])
        X_tr,y_tr = _create_seq(X_tr,y_raw,seq); X_te=X_te.reshape(1,seq,-1)

        def objective(trial):
            units = trial.suggest_int("units",16,128,log=True)
            drop  = trial.suggest_float("drop",0,0.4)
            lr    = trial.suggest_float("lr",1e-4,5e-3,log=True)
            m=keras.Sequential([
                layers.Input((seq,X_tr.shape[2])),
                layers.LSTM(units,dropout=drop),
                layers.Dense(1)
            ])
            m.compile(keras.optimizers.Adam(lr),loss='mse')
            es=callbacks.EarlyStopping(patience=EARLY_STOP,restore_best_weights=True,verbose=0)
            m.fit(X_tr,y_tr,epochs=200,batch_size=32,verbose=0,callbacks=[es])
            val=m.predict(X_tr[-int(0.1*len(X_tr)):],verbose=0).flatten()
            tru=y_tr[-int(0.1*len(y_tr)):]
            return np.sqrt(((val-tru)**2).mean())

        study=optuna.create_study(direction="minimize")
        study.optimize(objective,n_trials=15,show_progress_bar=False)
        u,d,lr=study.best_params.values()

        model=keras.Sequential([
            layers.Input((seq,X_tr.shape[2])),
            layers.LSTM(u,dropout=d),
            layers.Dense(1)
        ])
        model.compile(keras.optimizers.Adam(lr),loss='mse')
        es=callbacks.EarlyStopping(patience=EARLY_STOP,restore_best_weights=True,verbose=0)
        model.fit(X_tr,y_tr,epochs=200,batch_size=32,verbose=0,callbacks=[es])

        pred=float(model.predict(X_te,verbose=0)[0,0]); actual=float(targ[i])
        side=1 if pred>=0 else -1; sigma=targ[i-win:i].std(ddof=0)
        size=min(abs(pred)/sigma if sigma>0 else 0, MAX_LEV)
        pnl=max(side*size*actual, STOP_PCT)
        logs.append({
            "DataDate":dates[i].strftime('%y%m%d'),
            "Date":dates[i].strftime('%Y-%m-%d'),
            "PredRet":pred,"Actual":actual,"Size":size,"Side":side,
            "Err":pred-actual,"SE":(pred-actual)**2,"PnL":pnl,
            "Hit":int(np.sign(pred)==np.sign(actual))
        })
    return pd.DataFrame(logs)

# ── MAIN ────────────────────────────────────────────────────────────────────
def main():
    raw=pd.read_excel(IN_FILE,header=None,skiprows=2).iloc[:,:6]
    raw.columns=["Date","Open","High","Low","Close","Volume"]; raw=raw.drop("Open",axis=1)
    raw=raw.iloc[::-1].reset_index(drop=True)
    date=_parse_date; dates=raw["Date"].apply(date)
    price=_to_num(raw["Close"])
    valid=dates.notna()&price.notna(); dates=dates[valid]; price=price[valid]
    high=_to_num(raw["High"])[valid]; low=_to_num(raw["Low"])[valid]; vol=_to_num(raw["Volume"])[valid]
    ret=price.pct_change().mul(100)

    best=None; best_sh=-np.inf
    total = len(CAND_TRAIN) * len(CAND_SEQ)
    with tqdm(total=total, desc="Grid Search", unit="comb") as pbar:
        for w in CAND_TRAIN:
            for s in CAND_SEQ:
                feats=_tech_indicators(high,low,price,vol,lb=max(14,s)).dropna()
                df=pd.concat([dates.rename('Date'),price.rename('Close'),ret.rename('Ret'),feats],axis=1).dropna()
                if len(df)<w+s+10:
                    pbar.update(1)
                    continue
                wf=_walk(df,w,s)
                if wf.empty:
                    pbar.update(1)
                    continue
                sh=_pnl_stats(wf['PnL'])['Sharpe']
                if sh>best_sh: best_sh, best=(sh,(w,s,wf,df))
                pbar.update(1)

    if best is None:
        print("No valid parameter set."); return
    WIN,SEQ,wf,df=best[1]
    rmse=math.sqrt(wf['SE'].mean())
    stats=_pnl_stats(wf['PnL'])
    start=1_000_000; end=start*np.prod(1+wf['PnL']/100); yrs=len(wf)/252
    cagr=(end/start)**(1/yrs)-1; total=(end/start-1)*100

    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import Font
    wb=Workbook(); wb.remove(wb.active)
    def sheet(df,name):
        ws=wb.create_sheet(name)
        for r in dataframe_to_rows(df,index=False,header=True): ws.append(r)
        for r in ws.iter_rows(): 
            for c in r: c.font=Font(name="Times New Roman",size=12)
    sheet(wf,"WF_Log")
    ws=wb["WF_Log"]; ws.insert_rows(1); ws["A1"]="RMSE:"; ws["B1"]=rmse
    ws["A1"].font=ws["B1"].font=Font(name="Times New Roman",size=12,bold=True)
    kpi=pd.DataFrame({
        "Sharpe":[stats['Sharpe']],"CAGR":[cagr],"TotalPnL(%)":[total],
        "MaxDrawdown%": [stats['MaxDrawdown']],
        "Start_$":[f"{start:,.0f}"],"End_$":[f"{end:,.0f}"],
        "BestWin":[WIN],"BestSeqLen":[SEQ]})
    sheet(kpi,"Summary_PnL")

    tmp=Path(tempfile.gettempdir())
    imgs=_plot_yearly(wf[['Date','Actual','PredRet']],tmp)
    ws=wb.create_sheet("VS_Plots"); row=2
    for p in imgs:
        im=XLImage(str(p)); im.anchor=f"B{row}"; ws.add_image(im); row+=25
    for r in ws.iter_rows():
        for c in r: c.font=Font(name="Times New Roman",size=12)

    wb.save(OUT_FILE)
    print(f"Workbook saved → {OUT_FILE}")

if __name__=="__main__":
    t0=time.time(); main(); print(f"Elapsed: {time.time()-t0:.1f}s")